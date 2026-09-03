"""
Análise estrutural das 9 planilhas Cora
"""
import sys
from openpyxl import load_workbook

sheets_info = {}

files = [
    ('Fluxo de Caixa',          r'C:\Users\joaop\.minimax\v2\assets\2026\08\28\11-33-08-825-asset_20260828-113308-825_601eed269271_e3661db9-Planilha-de-Fluxo-de-Caixa-Cora-.xlsx'),
    ('Custo de Produção',       r'C:\Users\joaop\.minimax\v2\assets\2026\08\28\11-33-08-833-asset_20260828-113308-833_84bb4e917fad_3efc2ca5-Planilha-de-Custo-de-Producao-Cora.xlsx'),
    ('Conciliação Bancária',    r'C:\Users\joaop\.minimax\v2\assets\2026\08\28\11-33-08-838-asset_20260828-113308-838_5cbea5a822d1_20512a08-Planilha-de-Conciliacao-Bancaria-Cora.xlsx'),
    ('Precificação',            r'C:\Users\joaop\.minimax\v2\assets\2026\08\28\11-33-08-843-asset_20260828-113308-843_fef2724a3752_054d3fdd-Planilha-Cora-de-Precificacao.xlsx'),
    ('Contas a Pagar e Receber', r'C:\Users\joaop\.minimax\v2\assets\2026\08\28\11-33-08-848-asset_20260828-113308-848_3d1982208197_779f406a-Planilha-Cora-de-Contas-a-Pagar-e-Receber.xlsx'),
    ('Controle de Inadimplência', r'C:\Users\joaop\.minimax\v2\assets\2026\08\28\11-33-08-853-asset_20260828-113308-853_e1e52c0f247a_64396d54-Planilha-de-Controle-de-Inadimplencia-Cora.xlsx'),
    ('Balanço Patrimonial',     r'C:\Users\joaop\.minimax\v2\assets\2026\08\28\11-33-08-858-asset_20260828-113308-858_f45abb5b6f90_a91d73f8-Planilha-de-balanco-patrimonial-Cora-.xlsx'),
    ('Orçamento',                r'C:\Users\joaop\.minimax\v2\assets\2026\08\28\11-33-08-865-asset_20260828-113308-865_d348d9f61d22_d0f85928-Planilha-Cora-de-Orcamento-.xlsx'),
    ('Controle Financeiro Empresarial', r'C:\Users\joaop\.minimax\v2\assets\2026\08\28\11-33-08-871-asset_20260828-113308-871_bc7e92311615_64fd282b-Planilha-de-controle-financeiro-empresarial-Cora-.xlsx'),
]

for name, path in files:
    print(f'\n══════════════════════════════════════════════════════════')
    print(f'  {name}')
    print(f'══════════════════════════════════════════════════════════')
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        print(f'Sheets: {wb.sheetnames}')
        for sn in wb.sheetnames:
            ws = wb[sn]
            print(f'\n  --- Sheet: {sn} (max_row={ws.max_row}, max_col={ws.max_column}) ---')
            # Print first 5 rows
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 8: break
                vals = [str(c)[:40] if c is not None else '' for c in row]
                print(f'    R{i+1}: {vals[:10]}')
    except Exception as e:
        print(f'ERROR: {e}')
